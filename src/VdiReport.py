"""VDI 접속 로그를 읽어 Excel 보고서를 만드는 데스크톱 GUI 프로그램입니다.

이 프로그램이 하는 일
1. 사용자가 VDI 로그 Excel 파일(기본값: C:/VDI/1.vdilog.xlsx)을 선택합니다.
2. 사용자가 사용자정보 Excel 파일(기본값: C:/VDI/2.userinfo.xlsx)을 선택합니다.
3. 프로그램은 VDI 로그의 body 컬럼에서 사번과 접속 단말 IP를 찾아냅니다.
4. 프로그램은 사용자정보 파일에서 같은 사번을 찾아 이름, 이메일, 전화번호, 확인담당자 등을 붙입니다.
5. 합쳐진 결과를 화면 표(Treeview)에 미리 보여 줍니다.
6. 사용자는 미리보기 표의 마지막 열인 "접속사유"를 직접 입력하거나 수정할 수 있습니다.
7. 저장 버튼을 누르면 템플릿 Excel 파일(기본값: C:/VDI/3.VdiConnectReport.xlsx)에 데이터를 채워
   새 보고서 파일로 저장합니다.

초보자를 위한 큰 흐름
- "상수 영역": 기본 파일 경로, 보고서 헤더, 색상, 정규식처럼 프로그램 전체에서 쓰는 값을 모아 둡니다.
- "데이터 자료형": UserInfo, ReportRow처럼 한 사람/한 행의 데이터를 담는 작은 상자를 정의합니다.
- "데이터 처리 함수": Excel 읽기, 사번/IP 추출, 시간 변환, 보고서 행 만들기를 담당합니다.
- "Excel 저장 함수": 템플릿 파일을 열고, 행 서식을 복사하고, 테두리/정렬을 적용한 뒤 저장합니다.
- "VdiWindow 클래스": Tkinter 화면을 만들고 버튼 클릭, 표 정렬, 접속사유 편집 같은 사용자 동작을 처리합니다.
- "main 함수": Tkinter 창을 띄우고 프로그램 이벤트 루프를 시작합니다.

입력 파일 구조 가정
- VDI 로그 파일은 첫 번째 행을 헤더로 사용하며, 최소한 logtime, body 헤더가 필요합니다.
- 사용자정보 파일은 3행부터 실제 데이터가 있다고 가정합니다.
- 사용자정보 파일의 열 위치는 역할=B열, 이름=C열, 사번=D열, 전화번호=E열, 이메일=F열, 확인담당자=G열입니다.

주의할 점
- 이 파일은 GUI 프로그램이므로 main()을 실행하면 Tkinter 창이 열립니다.
- openpyxl은 Excel 파일을 읽고 쓰는 라이브러리이며, Excel 프로그램 자체를 실행하지는 않습니다.
- 사번과 IP는 로그 본문 텍스트에서 정규식으로 찾기 때문에 로그 형식이 크게 바뀌면 정규식을 수정해야 합니다.
"""

from __future__ import annotations

import platform
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import PhotoImage, StringVar, Tk
from tkinter import filedialog, messagebox, ttk
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


# 현재 실행 중인 운영체제를 한 번만 확인해 두고, 기본 경로/폰트를 정할 때 재사용합니다.
SYSTEM_NAME = platform.system()


def get_default_base_dir() -> Path:
    """프로그램에서 기본으로 사용할 VDI 폴더 위치를 운영체제에 맞게 정합니다."""
    if SYSTEM_NAME == "Windows":
        return Path("C:/VDI")
    return Path.home() / "VDI"


def get_default_ui_font() -> str:
    """한글이 깨지지 않도록 운영체제별 기본 UI 폰트를 선택합니다."""
    if SYSTEM_NAME == "Windows":
        return "Malgun Gothic"
    if SYSTEM_NAME == "Darwin":
        return "Apple SD Gothic Neo"
    return "Noto Sans CJK KR"


# 프로그램이 처음 열릴 때 자동으로 입력되는 기본 파일/폴더 경로입니다.
DEFAULT_BASE_DIR = get_default_base_dir()
DEFAULT_UI_FONT = get_default_ui_font()
DEFAULT_VDI_LOG_PATH = DEFAULT_BASE_DIR / "1.vdilog.xlsx"
DEFAULT_USER_INFO_PATH = DEFAULT_BASE_DIR / "2.userinfo.xlsx"
DEFAULT_OUTPUT_DIR = DEFAULT_BASE_DIR
TEMPLATE_PATH = DEFAULT_BASE_DIR / "3.VdiConnectReport.xlsx"

TITLE_TEXT = "VDI 접속현황"
WINDOW_TITLE = "VDI 접속 현황"
OUTPUT_FILENAME_PREFIX = "3.VdiConnectReport"

# 화면 전체에서 같이 사용할 색상입니다.
# 차분한 네이비/블루 계열을 사용해 업무용 시스템 도구처럼 보이도록 맞춥니다.
COLOR_APP_BG = "#EEF3F8"
COLOR_PANEL_BG = "#FFFFFF"
COLOR_HEADER_BG = "#12304A"
COLOR_HEADER_ACCENT = "#2D8CFF"
COLOR_TEXT = "#1D2B3A"
COLOR_MUTED_TEXT = "#627386"
COLOR_BORDER = "#D7E0EA"
COLOR_TABLE_HEADER = "#E7EEF6"
COLOR_TABLE_ODD = "#FFFFFF"
COLOR_TABLE_EVEN = "#F7FAFD"
COLOR_PRIMARY = "#1F6FEB"
COLOR_PRIMARY_ACTIVE = "#1557BA"
COLOR_SUCCESS = "#238636"
COLOR_SUCCESS_ACTIVE = "#1A6B2C"

# 보고서 제목과 Excel 헤더 이름을 상수로 모아 둡니다.
# 이렇게 해두면 화면 미리보기와 Excel 저장 시 같은 이름을 안전하게 재사용할 수 있습니다.
COL_NO = "No"
COL_CONNECTED_AT = "접속시간(KST)"
COL_CLIENT_IP = "접속단말 IP"
COL_ROLE = "역할"
COL_EMPLOYEE_NO = "사번"
COL_NAME = "이름"
COL_EMAIL = "이메일"
COL_PHONE = "전화번호"
COL_APPROVER = "확인담당자"
COL_REASON = "접속사유"

REPORT_HEADERS = [
    COL_NO,
    COL_CONNECTED_AT,
    COL_CLIENT_IP,
    COL_ROLE,
    COL_EMPLOYEE_NO,
    COL_NAME,
    COL_EMAIL,
    COL_PHONE,
    COL_APPROVER,
    COL_REASON,
]

# 로그 시간은 한국 시간 기준으로 보여주기 위해 KST 타임존을 사용합니다.
KST = ZoneInfo("Asia/Seoul")

# 로그 본문에서 사번과 IP를 찾기 위한 정규식입니다.
# EMPLOYEE_PATTERN은 "\+nb12345" 같은 형태에서 숫자 부분을 뽑아 "NB12345"로 맞춥니다.
EMPLOYEE_PATTERN = re.compile(r"(?i)\\+[n]?[b](\d{5,8})")
CLIENT_IP_PATTERN = re.compile(
    r'(?i)(?:ClientIP|Client IP|clientip|client_ip|ClientIpAddress|ForwardedClientIpAddress)\s*["=:]+\s*"?(?P<ip>[0-9]{1,3}(?:\.[0-9]{1,3}){3})'
)
IP_FALLBACK_PATTERN = re.compile(r"\b([0-9]{1,3}(?:\.[0-9]{1,3}){3})\b")

# Excel 데이터 영역에 적용할 기본 테두리 스타일입니다.
SOLID_SIDE = Side(style="thin", color="000000")
SOLID_BORDER = Border(
    left=SOLID_SIDE,
    right=SOLID_SIDE,
    top=SOLID_SIDE,
    bottom=SOLID_SIDE,
)


@dataclass
class UserInfo:
    """사용자정보 Excel에서 읽어 온 한 사람의 정보를 담는 자료형입니다."""

    # dataclass를 쓰면 __init__ 같은 기본 코드를 직접 만들지 않아도 됩니다.
    # 예: UserInfo(name="홍길동", employee_no="NB12345")처럼 쉽게 객체를 만들 수 있습니다.
    role: str = ""
    employee_no: str = ""
    name: str = ""
    email: str = ""
    phone: str = ""
    approver: str = ""


@dataclass
class ReportRow:
    """최종 보고서 한 줄에 들어갈 정보를 담는 자료형입니다."""

    # connected_at은 정렬용 datetime이고, connected_at_text는 화면/Excel 표시용 문자열입니다.
    # 이렇게 둘을 나누면 정렬은 정확하게 하고, 표시 형식은 보기 좋게 유지할 수 있습니다.
    connected_at: datetime
    connected_at_text: str
    client_ip: str
    role: str
    employee_no: str
    name: str
    email: str
    phone: str
    approver: str
    reason: str = ""

    def as_excel_row(self, number: int) -> list[Any]:
        """ReportRow 객체를 Excel 한 행에 바로 쓸 수 있는 리스트 형태로 바꿉니다."""
        return [
            number,
            self.connected_at_text,
            self.client_ip,
            self.role,
            self.employee_no,
            self.name,
            self.email,
            self.phone,
            self.approver,
            self.reason,
        ]


def value_to_text(value: Any) -> str:
    """Excel 셀 값처럼 여러 타입으로 들어올 수 있는 값을 안전한 문자열로 변환합니다."""
    # Excel 셀 값은 None, 숫자, 날짜, 문자열 등 여러 타입으로 들어올 수 있습니다.
    # 이후 정규식 검색이나 비교를 쉽게 하려면 먼저 문자열로 통일하는 것이 안전합니다.
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).replace("\u3000", "").strip()


def normalize_employee_no(value: Any) -> str:
    """사번 표기를 NB12345 형태로 통일합니다."""
    # 같은 사람의 사번이 파일마다 "B12345", "NB12345", "\+nb12345"처럼 다르면 매칭이 실패합니다.
    # 그래서 비교하기 전에 가능한 형태를 하나의 표준 형태인 "NB + 숫자"로 맞춥니다.
    text = value_to_text(value).upper().replace(" ", "")
    if not text:
        return ""

    # 로그나 사용자 정보에 "\+b12345", "\+nb12345"처럼 들어온 값도 같은 사번으로 인식합니다.
    match = EMPLOYEE_PATTERN.search(text)
    if match:
        return f"NB{match.group(1)}"

    # 이미 NB로 시작하면 그대로 쓰고, B로만 시작하면 앞에 N을 붙여 통일합니다.
    if re.fullmatch(r"NB\d{5,8}", text):
        return text
    if re.fullmatch(r"B\d{5,8}", text):
        return f"NB{text[1:]}"
    return text


def parse_logtime(value: Any) -> tuple[datetime, str]:
    """로그 시간 값을 datetime 객체와 화면 표시용 문자열로 변환합니다."""
    # 반환값이 2개인 이유:
    # 1. datetime 객체: 시간순 정렬에 사용합니다.
    # 2. 문자열: 미리보기 표와 Excel 셀에 표시합니다.
    if isinstance(value, datetime):
        # 시간대 정보가 있으면 KST로 바꾸고, Excel에 쓰기 쉽게 timezone 정보는 제거합니다.
        parsed = value.astimezone(KST).replace(tzinfo=None) if value.tzinfo else value
        return parsed, parsed.strftime("%Y-%m-%d %H:%M:%S")

    text = value_to_text(value)
    if not text:
        raise ValueError("logtime 값이 비어 있습니다.")

    candidate = text.replace("T", " ").replace("Z", "+00:00")
    parsed: datetime | None = None

    # ISO 형식(예: 2026-04-23T09:00:00Z)을 먼저 시도합니다.
    try:
        parsed = datetime.fromisoformat(candidate)
    except ValueError:
        # ISO 형식이 아니면 자주 쓰는 날짜/시간 형식을 하나씩 시도합니다.
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y/%m/%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M",
        ):
            try:
                parsed = datetime.strptime(text, fmt)
                break
            except ValueError:
                continue

    if parsed is None:
        raise ValueError(f"logtime 형식을 해석할 수 없습니다: {text}")

    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(KST).replace(tzinfo=None)

    return parsed, parsed.strftime("%Y-%m-%d %H:%M:%S")


def extract_employee_no(body_text: str) -> str:
    """로그 본문(body)에서 사번을 찾아 NB12345 형태로 반환합니다."""
    match = EMPLOYEE_PATTERN.search(body_text)
    if not match:
        return ""
    return f"NB{match.group(1)}"


def extract_client_ip(body_text: str) -> str:
    """로그 본문(body)에서 접속한 단말의 IP 주소를 찾습니다."""
    match = CLIENT_IP_PATTERN.search(body_text)
    if match:
        return match.group("ip")

    # ClientIP 같은 명확한 키가 없으면 본문에 있는 첫 번째 일반 IP를 보조로 사용합니다.
    for candidate in IP_FALLBACK_PATTERN.findall(body_text):
        if not candidate.startswith("127."):
            return candidate
    return ""


def load_log_rows(path: Path) -> list[dict[str, Any]]:
    """VDI 로그 Excel을 읽어, 각 행을 {'헤더명': 값} 형태의 딕셔너리로 변환합니다."""
    # data_only=True는 수식이 들어 있는 셀에서 수식 자체가 아니라 계산된 값을 읽도록 합니다.
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # 헤더는 소문자로 바꿔 둡니다. 그러면 "Body", "body", "BODY"처럼 대소문자가 달라도 row.get("body")로 찾을 수 있습니다.
    headers = [value_to_text(cell).lower() for cell in rows[0]]
    data_rows: list[dict[str, Any]] = []

    # 첫 행은 헤더이므로 두 번째 행부터 실제 데이터로 처리합니다.
    for row in rows[1:]:
        if all(value_to_text(cell) == "" for cell in row):
            continue

        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            # 행 길이가 헤더보다 짧을 수 있으므로 범위를 확인한 뒤 값을 넣습니다.
            item[header] = row[index] if index < len(row) else None
        data_rows.append(item)

    return data_rows


def load_user_info(path: Path) -> dict[str, UserInfo]:
    """사용자정보 Excel을 읽어 사번을 키로 하는 사용자 정보 사전을 만듭니다."""
    # 반환값을 dict로 만드는 이유:
    # 로그에서 사번을 찾은 뒤 users["NB12345"]처럼 빠르게 사용자 정보를 찾기 위해서입니다.
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    users: dict[str, UserInfo] = {}

    # 사용자정보 파일은 1~2행이 제목/헤더라고 보고, 3행부터 실제 사용자 데이터로 읽습니다.
    for row_index in range(3, sheet.max_row + 1):
        role = value_to_text(sheet.cell(row=row_index, column=2).value)
        name = value_to_text(sheet.cell(row=row_index, column=3).value)
        employee_no = normalize_employee_no(sheet.cell(row=row_index, column=4).value)
        phone = value_to_text(sheet.cell(row=row_index, column=5).value)
        email = value_to_text(sheet.cell(row=row_index, column=6).value)
        approver = value_to_text(sheet.cell(row=row_index, column=7).value)

        # 사번이 없는 행은 누구의 정보인지 알 수 없으므로 건너뜁니다.
        if not employee_no:
            continue

        users[employee_no] = UserInfo(
            role=role,
            employee_no=employee_no,
            name=name,
            email=email,
            phone=phone,
            approver=approver,
        )

    return users


def build_report_rows(log_path: Path, users: dict[str, UserInfo]) -> list[ReportRow]:
    """로그와 사용자 정보를 합쳐 최종 보고서에 들어갈 행 목록을 만듭니다."""
    # 이 함수가 실제 "보고서 데이터 만들기"의 중심입니다.
    # 로그 한 줄을 읽고 -> body에서 사번/IP를 찾고 -> 사용자정보와 합쳐 -> ReportRow로 저장합니다.
    report_rows: list[ReportRow] = []

    for row in load_log_rows(log_path):
        body = value_to_text(row.get("body"))
        if not body:
            continue

        # 로그 본문에서 사번을 찾지 못하면 사용자와 매칭할 수 없어 제외합니다.
        employee_no = extract_employee_no(body)
        if not employee_no:
            continue

        connected_at, connected_at_text = parse_logtime(row.get("logtime"))
        client_ip = extract_client_ip(body)

        # 사용자정보 파일에 없는 사번이어도 보고서에는 사번만이라도 남길 수 있게 기본값을 만듭니다.
        user = users.get(employee_no, UserInfo(employee_no=employee_no))

        report_rows.append(
            ReportRow(
                connected_at=connected_at,
                connected_at_text=connected_at_text,
                client_ip=client_ip,
                role=user.role,
                employee_no=employee_no,
                name=user.name,
                email=user.email,
                phone=user.phone,
                approver=user.approver,
                reason="",
            )
        )

    # 보고서가 시간 순서대로 보이도록 접속시간 기준 오름차순 정렬합니다.
    report_rows.sort(key=lambda item: item.connected_at)
    return report_rows


def clear_existing_data_rows(worksheet: Any) -> None:
    """템플릿에 남아 있을 수 있는 기존 데이터 행을 지웁니다."""
    if worksheet.max_row > 2:
        worksheet.delete_rows(3, worksheet.max_row - 2)


def copy_row_style(worksheet: Any, target_row: int, style_row: int, max_col: int) -> None:
    """템플릿의 예시 행 서식을 새 데이터 행에 복사합니다."""
    # 보고서 템플릿에는 보통 글꼴, 배경색, 테두리, 정렬 같은 서식이 미리 들어 있습니다.
    # 새 행에 값만 넣으면 서식이 깨질 수 있으므로, style_row의 서식을 target_row에 복사합니다.
    source_height = worksheet.row_dimensions[style_row].height
    if source_height is not None:
        worksheet.row_dimensions[target_row].height = source_height

    for column_index in range(1, max_col + 1):
        source_cell = worksheet.cell(row=style_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        target_cell._style = copy(source_cell._style)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)
            target_cell.number_format = source_cell.number_format


def apply_solid_border_to_data_area(worksheet: Any, start_row: int, end_row: int, max_col: int) -> None:
    """보고서 데이터 영역 전체에 얇은 검은색 테두리를 적용합니다."""
    for row_index in range(start_row, end_row + 1):
        for column_index in range(1, max_col + 1):
            worksheet.cell(row=row_index, column=column_index).border = SOLID_BORDER


def apply_data_alignment(worksheet: Any, start_row: int, end_row: int, max_col: int) -> None:
    """보고서 데이터 영역의 셀 정렬을 열 성격에 맞게 조정합니다."""
    for row_index in range(start_row, end_row + 1):
        for column_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            current_alignment = cell.alignment

            # No는 오른쪽, 접속사유는 왼쪽, 나머지는 가운데 정렬합니다.
            if column_index == 1:
                horizontal = "right"
            elif column_index == max_col:
                horizontal = "left"
            else:
                horizontal = "center"

            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical="center",
                text_rotation=getattr(current_alignment, "textRotation", 0),
                wrap_text=getattr(current_alignment, "wrap_text", None),
                shrink_to_fit=getattr(current_alignment, "shrink_to_fit", None),
                indent=getattr(current_alignment, "indent", 0),
                relativeIndent=getattr(current_alignment, "relativeIndent", None),
                justifyLastLine=getattr(current_alignment, "justifyLastLine", None),
                readingOrder=getattr(current_alignment, "readingOrder", None),
            )


def write_report_from_template(template_path: Path, output_path: Path, rows: list[ReportRow]) -> None:
    """템플릿 Excel 파일을 열어 보고서 데이터를 채운 뒤 새 파일로 저장합니다."""
    # 저장 흐름:
    # 1. 템플릿 파일을 엽니다.
    # 2. 예전에 남아 있던 데이터 행을 지웁니다.
    # 3. 제목/헤더를 다시 씁니다.
    # 4. rows 데이터를 3행부터 차례대로 씁니다.
    # 5. 테두리/정렬을 적용하고 output_path에 저장합니다.
    workbook = load_workbook(template_path)
    sheet = workbook.active
    max_col = len(REPORT_HEADERS)

    # 3행이 있으면 데이터 예시 행으로 보고 서식을 복사하고, 없으면 헤더 행 서식을 사용합니다.
    style_row = 3 if sheet.max_row >= 3 else 2

    clear_existing_data_rows(sheet)

    # 제목과 헤더는 항상 코드에 정의된 값으로 다시 써서 템플릿과 미리보기를 맞춥니다.
    sheet.cell(row=1, column=1, value=TITLE_TEXT)
    for column_index, header in enumerate(REPORT_HEADERS, start=1):
        sheet.cell(row=2, column=column_index, value=header)

    # Excel에서 1~2행은 제목/헤더라서 실제 데이터는 3행부터 씁니다.
    for index, row in enumerate(rows, start=1):
        target_row = index + 2
        copy_row_style(sheet, target_row, style_row, max_col)
        for column_index, value in enumerate(row.as_excel_row(index), start=1):
            sheet.cell(row=target_row, column=column_index, value=value)

    if rows:
        apply_solid_border_to_data_area(sheet, 3, len(rows) + 2, max_col)
        apply_data_alignment(sheet, 3, len(rows) + 2, max_col)

    workbook.save(output_path)


def validate_required_files(log_path: Path, user_info_path: Path) -> None:
    """보고서 생성에 필요한 입력 파일과 템플릿 파일이 모두 있는지 확인합니다."""
    # 미리보기를 만들거나 저장하기 전에 파일 존재 여부를 먼저 확인하면,
    # openpyxl에서 발생하는 복잡한 오류 대신 사용자에게 이해하기 쉬운 메시지를 보여줄 수 있습니다.
    required_files = (log_path, user_info_path, TEMPLATE_PATH)
    missing_files = [str(path) for path in required_files if not path.exists()]
    if missing_files:
        raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {', '.join(missing_files)}")


class VdiWindow:
    """VDI 보고서 생성 프로그램의 Tkinter 화면과 사용자 동작을 담당하는 클래스입니다.

    이 클래스 안에 들어 있는 것
    - __init__: 창 크기, 상태 변수, 미리보기 데이터 저장 공간을 준비합니다.
    - _configure_style: 버튼, 표, 라벨 등 화면 디자인을 설정합니다.
    - _build_ui: 실제 화면 위젯을 배치합니다.
    - select_* 함수: 파일/폴더 선택 창을 열고 선택값을 입력칸에 넣습니다.
    - generate_preview: Excel 파일을 읽어 미리보기 표를 만듭니다.
    - save_report: 미리보기 데이터를 Excel 보고서로 저장합니다.
    - open/close_reason_editor: 표의 접속사유 셀을 직접 편집할 수 있게 합니다.
    - sort_preview_by_column: 표 헤더를 클릭했을 때 정렬합니다.
    """

    def __init__(self, root: Tk) -> None:
        # root는 Tkinter가 제공하는 가장 바깥쪽 창 객체입니다.
        # 이 객체 위에 라벨, 버튼, 표 같은 위젯을 하나씩 올려 화면을 구성합니다.
        self.root = root
        self.root.title(WINDOW_TITLE)
        self.root.geometry("1440x860")
        self.root.minsize(1100, 600)

        # 미리보기 표에 표시할 보고서 행과, 각 열의 정렬 방향을 기억합니다.
        self.preview_rows: list[ReportRow] = []
        self.sort_reverse: dict[str, bool] = {}

        # StringVar는 Tkinter 입력창/라벨 값과 Python 변수를 연결해 주는 객체입니다.
        self.log_path_var = StringVar(value=str(DEFAULT_VDI_LOG_PATH))
        self.user_info_path_var = StringVar(value=str(DEFAULT_USER_INFO_PATH))
        self.output_dir_var = StringVar(value=str(DEFAULT_OUTPUT_DIR))
        self.status_var = StringVar(
            value="VDI 로그 파일과 사용자정보 파일을 선택한 뒤 미리보기를 실행하세요."
        )

        # Treeview는 각 열에 내부 ID가 필요하므로 col_0, col_1 같은 이름을 만들어 둡니다.
        self.columns = [f"col_{index}" for index in range(len(REPORT_HEADERS))]
        self.tree: ttk.Treeview
        self.icons = self._build_icons()

        # 접속사유 열을 클릭했을 때 잠깐 나타나는 입력창과 편집 중인 행을 관리합니다.
        self.reason_editor: ttk.Entry | None = None
        self.reason_editor_var = StringVar()
        self.editing_item_id: str | None = None

        self._configure_style()
        self._build_ui()

    def _configure_style(self) -> None:
        """버튼/라벨 같은 기본 위젯 스타일을 설정합니다."""
        # ttk.Style은 Tkinter 위젯의 공통 디자인을 이름으로 등록하는 도구입니다.
        # 예를 들어 "Primary.TButton" 스타일을 만든 뒤 Button에 style="Primary.TButton"을 지정하면
        # 같은 색상/패딩/폰트를 여러 버튼에서 재사용할 수 있습니다.
        style = ttk.Style(self.root)
        try:
            # clam 테마는 Windows/macOS/Linux에서 색상 커스터마이징이 비교적 잘 적용됩니다.
            style.theme_use("clam")
        except Exception:
            pass

        self.root.configure(background=COLOR_APP_BG)

        style.configure(".", font=(DEFAULT_UI_FONT, 10), foreground=COLOR_TEXT)
        style.configure("App.TFrame", background=COLOR_APP_BG)
        style.configure("Hero.TFrame", background=COLOR_HEADER_BG)
        style.configure("Card.TFrame", background=COLOR_PANEL_BG, borderwidth=1, relief="solid")
        style.configure("Toolbar.TFrame", background=COLOR_PANEL_BG)
        style.configure("TableCard.TFrame", background=COLOR_PANEL_BG, borderwidth=1, relief="solid")

        style.configure(
            "HeroTitle.TLabel",
            background=COLOR_HEADER_BG,
            foreground="#FFFFFF",
            font=(DEFAULT_UI_FONT, 19, "bold"),
        )
        style.configure(
            "HeroSubtitle.TLabel",
            background=COLOR_HEADER_BG,
            foreground="#C8D7EA",
            font=(DEFAULT_UI_FONT, 10),
        )
        style.configure(
            "CardTitle.TLabel",
            background=COLOR_PANEL_BG,
            foreground=COLOR_TEXT,
            font=(DEFAULT_UI_FONT, 11, "bold"),
        )
        style.configure("Header.TLabel", background=COLOR_PANEL_BG, foreground=COLOR_MUTED_TEXT, font=(DEFAULT_UI_FONT, 10, "bold"))
        style.configure("Hint.TLabel", background=COLOR_PANEL_BG, foreground=COLOR_MUTED_TEXT, font=(DEFAULT_UI_FONT, 9))
        style.configure("Status.TLabel", background="#DDE7F2", foreground=COLOR_TEXT, padding=(12, 8))

        style.configure("Path.TEntry", fieldbackground="#F8FBFE", bordercolor=COLOR_BORDER, padding=(6, 4))
        style.configure("Action.TButton", padding=(14, 9), font=(DEFAULT_UI_FONT, 10, "bold"))
        style.configure("Primary.TButton", padding=(14, 9), font=(DEFAULT_UI_FONT, 10, "bold"), foreground="#FFFFFF", background=COLOR_PRIMARY)
        style.map("Primary.TButton", background=[("active", COLOR_PRIMARY_ACTIVE), ("pressed", COLOR_PRIMARY_ACTIVE)])
        style.configure("Success.TButton", padding=(14, 9), font=(DEFAULT_UI_FONT, 10, "bold"), foreground="#FFFFFF", background=COLOR_SUCCESS)
        style.map("Success.TButton", background=[("active", COLOR_SUCCESS_ACTIVE), ("pressed", COLOR_SUCCESS_ACTIVE)])
        style.configure("Secondary.TButton", padding=(14, 9), foreground=COLOR_TEXT, background="#E8EEF5")
        style.map("Secondary.TButton", background=[("active", "#D7E2EE"), ("pressed", "#D7E2EE")])
        style.configure("Browse.TButton", padding=(10, 7), foreground=COLOR_TEXT, background="#EDF3FA")

        style.configure(
            "Report.Treeview",
            background=COLOR_TABLE_ODD,
            fieldbackground=COLOR_TABLE_ODD,
            foreground=COLOR_TEXT,
            rowheight=31,
            bordercolor=COLOR_BORDER,
            borderwidth=1,
        )
        style.configure(
            "Report.Treeview.Heading",
            background=COLOR_TABLE_HEADER,
            foreground=COLOR_TEXT,
            font=(DEFAULT_UI_FONT, 10, "bold"),
            padding=(8, 7),
        )
        style.map(
            "Report.Treeview",
            background=[("selected", COLOR_HEADER_ACCENT)],
            foreground=[("selected", "#FFFFFF")],
        )

    def _build_ui(self) -> None:
        """파일 선택 영역, 버튼 영역, 미리보기 표, 상태바를 화면에 배치합니다."""
        # grid(row, column)은 화면을 표처럼 나누어 위젯을 배치하는 방식입니다.
        # sticky="ew"는 좌우로 늘리고, sticky="nsew"는 상하좌우로 늘리라는 뜻입니다.
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(3, weight=1)

        # 상단 배너: 프로그램의 목적과 현재 작업 흐름을 한눈에 보여줍니다.
        hero_frame = ttk.Frame(self.root, padding=(22, 18, 22, 18), style="Hero.TFrame")
        hero_frame.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 10))
        hero_frame.columnconfigure(0, weight=1)
        ttk.Label(hero_frame, text="VDI 접속 현황 보고서", style="HeroTitle.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            hero_frame,
            text="로그 Excel과 사용자정보 Excel을 병합해 검토 가능한 보고서 미리보기와 저장 파일을 생성합니다.",
            style="HeroSubtitle.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Label(
            hero_frame,
            text="1. 파일 선택  2. 미리보기 생성  3. 접속사유 확인  4. 보고서 저장",
            style="HeroSubtitle.TLabel",
        ).grid(row=2, column=0, sticky="w", pady=(8, 0))

        # 파일 카드: 로그 파일, 사용자정보 파일, 출력 폴더를 선택하는 영역입니다.
        file_frame = ttk.Frame(self.root, padding=(18, 16, 18, 16), style="Card.TFrame")
        file_frame.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        ttk.Label(file_frame, text="입력 및 출력 경로", style="CardTitle.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 10)
        )

        self._add_path_row(
            file_frame,
            row=1,
            label_text="VDI 로그파일",
            variable=self.log_path_var,
            button_text="찾아보기",
            command=self.select_log_file,
        )
        self._add_path_row(
            file_frame,
            row=2,
            label_text="사용자정보파일",
            variable=self.user_info_path_var,
            button_text="찾아보기",
            command=self.select_user_info_file,
        )
        self._add_path_row(
            file_frame,
            row=3,
            label_text="출력폴더",
            variable=self.output_dir_var,
            button_text="폴더선택",
            command=self.select_output_dir,
        )

        # 액션 바: 주요 실행 버튼을 한 줄에 배치합니다.
        button_frame = ttk.Frame(self.root, padding=(18, 12, 18, 12), style="Toolbar.TFrame")
        button_frame.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 10))
        ttk.Button(
            button_frame,
            text="미리보기 생성",
            image=self.icons["preview"],
            compound="left",
            style="Primary.TButton",
            command=self.generate_preview,
        ).pack(side="left", padx=(0, 10))
        ttk.Button(
            button_frame,
            text="파일 저장",
            image=self.icons["save"],
            compound="left",
            style="Success.TButton",
            command=self.save_report,
        ).pack(side="left", padx=(0, 10))
        ttk.Button(
            button_frame,
            text="미리보기 초기화",
            image=self.icons["clear"],
            compound="left",
            style="Secondary.TButton",
            command=self.clear_preview,
        ).pack(side="left")
        ttk.Label(
            button_frame,
            text="접속사유는 미리보기 표의 마지막 열을 클릭해 바로 수정할 수 있습니다.",
            style="Hint.TLabel",
        ).pack(side="right")

        # 데이터 카드: 보고서 내용을 표 형태로 미리 보여주는 Treeview입니다.
        table_frame = ttk.Frame(self.root, padding=(14, 12, 14, 14), style="TableCard.TFrame")
        table_frame.grid(row=3, column=0, sticky="nsew", padx=14, pady=(0, 10))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)
        ttk.Label(table_frame, text="보고서 미리보기", style="CardTitle.TLabel").grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 10)
        )

        self.tree = ttk.Treeview(
            table_frame,
            columns=self.columns,
            show="headings",
            selectmode="browse",
            height=20,
            style="Report.Treeview",
        )
        self.tree.grid(row=1, column=0, sticky="nsew")
        self.tree.bind("<ButtonRelease-1>", self.handle_tree_click)
        self.tree.bind("<Configure>", lambda _event: self.close_reason_editor(save=True))
        self.tree.tag_configure("odd", background=COLOR_TABLE_ODD)
        self.tree.tag_configure("even", background=COLOR_TABLE_EVEN)

        # 데이터가 많을 때 볼 수 있도록 세로/가로 스크롤바를 연결합니다.
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        vsb.grid(row=1, column=1, sticky="ns")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        hsb.grid(row=2, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        widths = [70, 170, 140, 110, 120, 120, 220, 130, 130, 220]
        anchors = ["e", "center", "center", "center", "center", "center", "center", "center", "center", "w"]

        # 각 열의 제목, 폭, 정렬을 설정합니다. 제목을 클릭하면 해당 열로 정렬됩니다.
        for index, (column_id, header, width, anchor) in enumerate(
            zip(self.columns, REPORT_HEADERS, widths, anchors)
        ):
            self.tree.heading(column_id, text=header, command=lambda idx=index: self.sort_preview_by_column(idx))
            self.tree.column(column_id, width=width, minwidth=80, anchor=anchor, stretch=True)

        status = ttk.Label(self.root, textvariable=self.status_var, style="Status.TLabel")
        status.grid(row=4, column=0, sticky="ew", padx=14, pady=(0, 14))

    def _add_path_row(
        self,
        parent: ttk.Frame,
        row: int,
        label_text: str,
        variable: StringVar,
        button_text: str,
        command: Any,
    ) -> None:
        """파일/폴더 경로를 입력하는 한 줄짜리 UI를 추가합니다."""
        ttk.Label(parent, text=label_text, style="Header.TLabel").grid(
            row=row, column=0, sticky="w", padx=(0, 12), pady=6
        )
        ttk.Entry(parent, textvariable=variable, style="Path.TEntry").grid(row=row, column=1, sticky="ew", pady=6)
        ttk.Button(parent, text=button_text, command=command, style="Browse.TButton").grid(
            row=row, column=2, sticky="ew", padx=(10, 0), pady=6
        )

    def _build_icons(self) -> dict[str, PhotoImage]:
        """버튼에 표시할 작은 아이콘 이미지를 코드로 직접 만듭니다."""
        return {
            "preview": self._create_preview_icon(),
            "save": self._create_save_icon(),
            "clear": self._create_clear_icon(),
        }

    def _get_existing_initial_dir(self, current_text: str, fallback: Path) -> Path:
        """파일 선택 창이 처음 열릴 폴더를 현재 입력값이나 기본값에서 안전하게 찾습니다."""
        current_path = Path(current_text).expanduser() if current_text.strip() else fallback
        if current_path.exists():
            return current_path if current_path.is_dir() else current_path.parent

        fallback = fallback.expanduser()
        if fallback.exists():
            return fallback if fallback.is_dir() else fallback.parent
        return Path.home()

    def _fill_rect(self, image: PhotoImage, x1: int, y1: int, x2: int, y2: int, color: str) -> None:
        """PhotoImage 위에 사각형을 칠하는 보조 함수입니다."""
        image.put(color, to=(x1, y1, x2, y2))

    def _draw_circle(self, image: PhotoImage, center_x: int, center_y: int, radius: int, color: str) -> None:
        """PhotoImage 위에 원을 픽셀 단위로 그리는 보조 함수입니다."""
        for y in range(center_y - radius, center_y + radius + 1):
            for x in range(center_x - radius, center_x + radius + 1):
                if (x - center_x) ** 2 + (y - center_y) ** 2 <= radius**2:
                    image.put(color, (x, y))

    def _create_preview_icon(self) -> PhotoImage:
        """미리보기 버튼용 아이콘을 만듭니다."""
        image = PhotoImage(width=24, height=24)
        self._fill_rect(image, 4, 2, 15, 20, "#3B7BD6")
        self._fill_rect(image, 6, 4, 13, 18, "#FFFFFF")
        self._fill_rect(image, 11, 2, 15, 6, "#9CC2F4")
        self._fill_rect(image, 7, 7, 12, 8, "#C8D7EE")
        self._fill_rect(image, 7, 10, 12, 11, "#C8D7EE")
        self._fill_rect(image, 7, 13, 11, 14, "#C8D7EE")
        self._draw_circle(image, 15, 15, 5, "#1DB9C3")
        self._draw_circle(image, 15, 15, 3, "#DDEFF8")
        self._fill_rect(image, 18, 18, 22, 22, "#28507E")
        self._fill_rect(image, 17, 17, 19, 19, "#28507E")
        return image

    def _create_save_icon(self) -> PhotoImage:
        """저장 버튼용 아이콘을 만듭니다."""
        image = PhotoImage(width=24, height=24)
        self._fill_rect(image, 3, 2, 16, 20, "#3B7BD6")
        self._fill_rect(image, 5, 4, 14, 18, "#FFFFFF")
        self._fill_rect(image, 6, 5, 13, 8, "#D7E4F6")
        self._fill_rect(image, 10, 5, 12, 8, "#285FA8")
        self._fill_rect(image, 7, 12, 12, 13, "#C8D7EE")
        self._fill_rect(image, 7, 15, 12, 16, "#C8D7EE")
        self._fill_rect(image, 17, 5, 21, 15, "#3FAE57")
        self._fill_rect(image, 15, 13, 23, 16, "#3FAE57")
        self._fill_rect(image, 18, 15, 20, 20, "#3FAE57")
        self._fill_rect(image, 16, 17, 22, 19, "#3FAE57")
        return image

    def _create_clear_icon(self) -> PhotoImage:
        """초기화 버튼용 아이콘을 만듭니다."""
        image = PhotoImage(width=24, height=24)
        self._fill_rect(image, 5, 3, 18, 20, "#425875")
        self._fill_rect(image, 7, 5, 16, 18, "#FFFFFF")
        self._fill_rect(image, 8, 8, 14, 9, "#CCD6E4")
        self._fill_rect(image, 8, 11, 14, 12, "#CCD6E4")
        self._fill_rect(image, 8, 14, 13, 15, "#CCD6E4")
        self._draw_circle(image, 7, 7, 5, "#FF8A1D")
        self._draw_circle(image, 7, 7, 3, "#FFFFFF")
        self._fill_rect(image, 8, 2, 12, 5, "#FF8A1D")
        self._fill_rect(image, 14, 14, 22, 20, "#6C809E")
        self._fill_rect(image, 14, 18, 22, 20, "#ECEFF4")
        return image

    def select_log_file(self) -> None:
        """VDI 로그 Excel 파일을 선택하고 입력칸에 경로를 넣습니다."""
        selected = filedialog.askopenfilename(
            title="VDI 로그파일 선택",
            initialdir=str(self._get_existing_initial_dir(self.log_path_var.get(), DEFAULT_VDI_LOG_PATH.parent)),
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if selected:
            self.log_path_var.set(selected)

    def select_user_info_file(self) -> None:
        """사용자정보 Excel 파일을 선택하고 입력칸에 경로를 넣습니다."""
        selected = filedialog.askopenfilename(
            title="사용자정보파일 선택",
            initialdir=str(self._get_existing_initial_dir(self.user_info_path_var.get(), DEFAULT_USER_INFO_PATH.parent)),
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")],
        )
        if selected:
            self.user_info_path_var.set(selected)

    def select_output_dir(self) -> None:
        """보고서를 저장할 출력 폴더를 선택하고 입력칸에 경로를 넣습니다."""
        selected = filedialog.askdirectory(
            title="출력폴더 선택",
            initialdir=str(self._get_existing_initial_dir(self.output_dir_var.get(), DEFAULT_OUTPUT_DIR)),
        )
        if selected:
            self.output_dir_var.set(selected)

    def clear_preview(self) -> None:
        """미리보기 표와 정렬 상태를 모두 초기화합니다."""
        self.close_reason_editor(save=False)
        self.preview_rows = []
        self.sort_reverse = {}
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)
        self.status_var.set("미리보기가 초기화되었습니다.")

    def generate_preview(self) -> None:
        """입력 Excel 파일들을 읽어 보고서 행을 만들고 미리보기 표에 표시합니다."""
        try:
            # 사용자가 접속사유를 편집 중이었다면, 미리보기를 다시 만들기 전에 현재 입력값을 먼저 저장합니다.
            self.close_reason_editor(save=True)
            log_path = Path(self.log_path_var.get().strip())
            user_info_path = Path(self.user_info_path_var.get().strip())

            # 필요한 파일이 모두 있는지 먼저 확인한 뒤 실제 데이터를 읽습니다.
            validate_required_files(log_path, user_info_path)

            # 사용자정보를 먼저 사전(dict)으로 읽어 둡니다.
            # 그다음 로그를 읽으면서 사번별 사용자정보를 빠르게 붙입니다.
            users = load_user_info(user_info_path)
            self.preview_rows = build_report_rows(log_path, users)

            # 새 미리보기를 만들면 이전 정렬 방향은 의미가 없으므로 초기화합니다.
            self.sort_reverse = {}
            self.refresh_table()
            self.status_var.set(f"미리보기 생성 완료: {len(self.preview_rows)}건")
        except Exception as exc:
            messagebox.showerror("오류", str(exc))
            self.status_var.set("미리보기 생성 중 오류가 발생했습니다.")

    def refresh_table(self) -> None:
        """preview_rows에 들어 있는 데이터를 Treeview 표에 다시 그립니다."""
        # Treeview는 화면에 보이는 표이고, preview_rows는 실제 데이터 목록입니다.
        # 정렬/수정 후에는 preview_rows 기준으로 Treeview를 비웠다가 다시 그려 화면과 데이터를 맞춥니다.
        self.close_reason_editor(save=True)
        for item_id in self.tree.get_children():
            self.tree.delete(item_id)

        for row_index, row in enumerate(self.preview_rows, start=1):
            values = row.as_excel_row(row_index)
            row_tag = "even" if row_index % 2 == 0 else "odd"
            self.tree.insert("", "end", iid=str(row_index - 1), values=values, tags=(row_tag,))

    def handle_tree_click(self, event: Any) -> None:
        """미리보기 표를 클릭했을 때 접속사유 열이면 편집창을 엽니다."""
        if self.tree.identify_region(event.x, event.y) != "cell":
            self.close_reason_editor(save=True)
            return

        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)

        # 마지막 열(접속사유)만 직접 편집할 수 있게 제한합니다.
        if not item_id or column_id != f"#{len(REPORT_HEADERS)}":
            self.close_reason_editor(save=True)
            return

        self.open_reason_editor(item_id, column_id)

    def open_reason_editor(self, item_id: str, column_id: str) -> None:
        """선택한 접속사유 셀 위에 Entry 입력창을 겹쳐 표시합니다."""
        # Treeview는 기본적으로 셀을 직접 편집하는 기능이 없습니다.
        # 그래서 클릭한 셀 위치(bbox)를 구한 뒤, 그 위치에 Entry 입력창을 잠깐 올려 편집처럼 보이게 합니다.
        self.close_reason_editor(save=True)
        bbox = self.tree.bbox(item_id, column_id)
        if not bbox:
            return

        x, y, width, height = bbox
        row = self.preview_rows[int(item_id)]
        self.editing_item_id = item_id
        self.reason_editor_var.set(row.reason)
        self.reason_editor = ttk.Entry(self.tree, textvariable=self.reason_editor_var)
        self.reason_editor.place(x=x, y=y, width=width, height=height)
        self.reason_editor.focus_set()
        self.reason_editor.selection_range(0, "end")
        self.reason_editor.bind("<Return>", lambda _event: self.close_reason_editor(save=True))
        self.reason_editor.bind("<Escape>", lambda _event: self.close_reason_editor(save=False))
        self.reason_editor.bind("<FocusOut>", lambda _event: self.close_reason_editor(save=True))

    def close_reason_editor(self, save: bool) -> None:
        """접속사유 편집창을 닫고, 필요하면 입력값을 preview_rows에 저장합니다."""
        if self.reason_editor is None:
            return

        item_id = self.editing_item_id
        editor = self.reason_editor
        value = self.reason_editor_var.get().strip()

        self.reason_editor = None
        self.editing_item_id = None
        editor.destroy()

        # Enter/포커스 이동은 저장하고, Escape/초기화는 저장하지 않도록 save 값으로 구분합니다.
        if save and item_id is not None and item_id.isdigit():
            row_index = int(item_id)
            if 0 <= row_index < len(self.preview_rows):
                self.preview_rows[row_index].reason = value
                values = self.preview_rows[row_index].as_excel_row(row_index + 1)
                self.tree.item(item_id, values=values)
                self.status_var.set(f"접속사유 수정 완료: {self.preview_rows[row_index].employee_no}")

    def sort_preview_by_column(self, column_index: int) -> None:
        """사용자가 표 헤더를 클릭하면 해당 열 기준으로 미리보기 데이터를 정렬합니다."""
        if not self.preview_rows:
            return

        # 정렬하기 전에 편집 중인 접속사유가 있으면 먼저 저장합니다.
        # 저장하지 않고 정렬하면 어떤 행을 수정 중이었는지 헷갈릴 수 있습니다.
        self.close_reason_editor(save=True)

        reverse = self.sort_reverse.get(self.columns[column_index], False)

        # 열마다 데이터 성격이 다르므로 날짜/IP/문자열에 맞는 기준으로 정렬합니다.
        if column_index == 0:
            self.preview_rows.reverse()
        elif column_index == 1:
            self.preview_rows.sort(key=lambda row: row.connected_at, reverse=reverse)
        elif column_index == 2:
            self.preview_rows.sort(
                key=lambda row: tuple(int(part) for part in row.client_ip.split(".")) if row.client_ip else (),
                reverse=reverse,
            )
        elif column_index == 3:
            self.preview_rows.sort(key=lambda row: row.role.casefold(), reverse=reverse)
        elif column_index == 4:
            self.preview_rows.sort(key=lambda row: row.employee_no.casefold(), reverse=reverse)
        elif column_index == 5:
            self.preview_rows.sort(key=lambda row: row.name.casefold(), reverse=reverse)
        elif column_index == 6:
            self.preview_rows.sort(key=lambda row: row.email.casefold(), reverse=reverse)
        elif column_index == 7:
            self.preview_rows.sort(key=lambda row: row.phone.casefold(), reverse=reverse)
        elif column_index == 8:
            self.preview_rows.sort(key=lambda row: row.approver.casefold(), reverse=reverse)
        elif column_index == 9:
            self.preview_rows.sort(key=lambda row: row.reason.casefold(), reverse=reverse)

        self.sort_reverse[self.columns[column_index]] = not reverse
        self.refresh_table()

        direction = "내림차순" if reverse else "오름차순"
        self.status_var.set(f"미리보기 정렬 완료: {REPORT_HEADERS[column_index]} {direction}")

    def build_output_path(self, suffix: str = ".xlsx") -> Path:
        """출력 폴더와 현재 시간을 이용해 중복 가능성이 낮은 저장 파일명을 만듭니다."""
        # 저장 파일명 예: 3.VdiConnectReport_20260423_1015.xlsx
        # 시간을 붙이면 같은 폴더에 여러 번 저장해도 기존 파일을 덮어쓸 가능성이 줄어듭니다.
        output_dir_text = self.output_dir_var.get().strip()
        if not output_dir_text:
            raise ValueError("출력폴더를 선택해 주세요.")

        output_dir = Path(output_dir_text)
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return output_dir / f"{OUTPUT_FILENAME_PREFIX}_{timestamp}{suffix}"

    def save_report(self) -> None:
        """미리보기 데이터를 템플릿에 써서 최종 Excel 보고서로 저장합니다."""
        try:
            # 저장 직전에 편집 중인 접속사유가 있다면 먼저 반영합니다.
            self.close_reason_editor(save=True)

            # 사용자가 미리보기를 만들지 않고 저장을 눌렀다면 먼저 미리보기를 자동 생성합니다.
            if not self.preview_rows:
                self.generate_preview()
                if not self.preview_rows:
                    return

            output_path = self.build_output_path()
            write_report_from_template(TEMPLATE_PATH, output_path, self.preview_rows)
            messagebox.showinfo("완료", f"보고서가 저장되었습니다.\n{output_path}")
            self.status_var.set(f"파일 저장 완료: {output_path}")
        except Exception as exc:
            messagebox.showerror("오류", str(exc))
            self.status_var.set("파일 저장 중 오류가 발생했습니다.")


def main() -> None:
    """Tkinter 프로그램을 시작하는 진입점입니다."""
    # Tk()는 GUI 프로그램의 메인 창을 만듭니다.
    root = Tk()

    # VdiWindow 객체를 만들면서 창 안에 버튼, 입력칸, 표 같은 위젯이 배치됩니다.
    VdiWindow(root)

    # mainloop()는 사용자의 클릭/키보드 입력을 기다리는 Tkinter 이벤트 루프입니다.
    # 이 줄이 실행되어야 창이 바로 닫히지 않고 계속 떠 있습니다.
    root.mainloop()


if __name__ == "__main__":
    main()
